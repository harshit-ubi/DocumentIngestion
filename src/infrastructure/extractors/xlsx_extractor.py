import io
import openpyxl
from src.domain.interfaces.extractor import DocumentExtractorInterface
from src.core.exceptions import ExtractionError
from src.core.logging import logger


class XlsxExtractor(DocumentExtractorInterface):
    """XLSX Excel Document Extractor using openpyxl."""

    def extract_text(self, file_content: bytes, filename: str) -> str:
        try:
            excel_file = io.BytesIO(file_content)
            workbook = openpyxl.load_workbook(excel_file, data_only=True)
            extracted_sheets = []

            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                rows_text = []

                for row in sheet.iter_rows(values_only=True):
                    # Filter out empty cells in row
                    row_values = [str(val).strip() for val in row if val is not None and str(val).strip() != ""]
                    if row_values:
                        rows_text.append(" | ".join(row_values))

                if rows_text:
                    extracted_sheets.append(f"=== Sheet: {sheet_name} ===\n" + "\n".join(rows_text))

            if not extracted_sheets:
                raise ExtractionError(f"No extractable text content found in XLSX '{filename}'.")

            logger.info(f"Successfully extracted {len(extracted_sheets)} sheets from XLSX '{filename}'.")
            return "\n\n".join(extracted_sheets)
        except Exception as e:
            logger.error(f"Failed to extract text from XLSX '{filename}': {str(e)}")
            raise ExtractionError(f"Failed to extract text from XLSX '{filename}': {str(e)}")
